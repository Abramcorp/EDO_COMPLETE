"""Добавляет 84pt (9pt + 75pt) пустых строк в конец ключевых листов xlsx-шаблонов.

Цель: зарезервировать в нижней части каждой страницы декларации место под
штамп edo-stamps, который оверлеем рисуется в зоне y=12..87pt от низа A4.

Принцип взят из эталона Тензора (Романов_УСН_2025.xml):
- Row N (h=9pt): резерв под текст "Оператор ЭДО ООО Тензор"
- Row N+1 (h=75pt): резерв под рамку "ДОКУМЕНТ ПОДПИСАН + Идентификатор"

После soffice fitToPage=True эти 84pt окажутся в нижних 84pt страницы A4.

ВАЖНО: openpyxl сохраняет row_dimensions только если в строке есть хотя бы
одна реальная ячейка. Поэтому в каждую новую строку пишем пробел в A, чтобы
строка попала в итоговый xml и max_row сдвинулся.
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parents[1] / 'modules' / 'usn_declaration' / 'data'

# Какие листы нужны нашему пайплайну для УСН "Доходы".
TARGETS = {
    'declaration_template_2024.xlsx': [
        'Титул',
        'Раздел 1.1',
        'Раздел 2.1.1',
        'Раздел 2.1.1 (продолжение)',
    ],
    'declaration_template_2025.xlsx': [
        'стр.1',
        'стр.2_Разд.1',
        'стр.3_Разд.2',
    ],
    'declaration_template.xlsx': [
        'Титульный лист',
        'Р.1.1',
        'Р.2.1.1',
        'Р.2.1.1 (продол.)',
    ],
}

PAD_HEIGHTS = [9.0, 75.0]  # pt — точно как у Тензора (rows 52-53)


def add_padding(ws) -> tuple[int, int, float]:
    """Добавляет PAD_HEIGHTS в конец листа.

    Чтобы строки реально попали в xlsx (а не только в row_dimensions),
    пишем пробел в ячейку A каждой новой строки.

    Возвращает (row_before, row_after, total_added_pt).
    """
    row_before = ws.max_row
    next_r = row_before + 1
    for h in PAD_HEIGHTS:
        ws.row_dimensions[next_r].height = h
        ws.cell(row=next_r, column=1, value=' ')  # пробел чтобы строка существовала
        next_r += 1
    return (row_before, ws.max_row, sum(PAD_HEIGHTS))


def main() -> int:
    if not DATA_DIR.is_dir():
        print(f'ERROR: data dir not found: {DATA_DIR}', file=sys.stderr)
        return 1

    for tpl_name, sheets in TARGETS.items():
        path = DATA_DIR / tpl_name
        if not path.is_file():
            print(f'SKIP: {tpl_name} (not found)')
            continue
        print(f'\n=== {tpl_name} ===')
        wb = load_workbook(path)
        for s in sheets:
            if s not in wb.sheetnames:
                print(f'  WARN: sheet {s!r} not in workbook')
                continue
            ws = wb[s]
            before, after, added = add_padding(ws)
            print(f'  "{s}": rows {before} -> {after} (+{int(added)}pt)')
        wb.save(path)
        print(f'  saved {path.name}')

    print('\nDone.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
