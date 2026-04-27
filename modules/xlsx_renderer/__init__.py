"""
Рендер декларации УСН через заполнение xlsx-шаблона ФНС с последующей
конвертацией в PDF через LibreOffice headless.

Алгоритм:
1. Загружаем templates/knd_1152017/declaration_template_2024.xlsx
2. Заполняем ячейки данными через openpyxl (_safe_set для merged cells)
3. Структурно модифицируем Титул — сжимаем пустые строки 30-39 и
   55-63 чтобы освободить зону y=30..180pt под штамп ЭДО
4. Удаляем ненужные листы (Р.1.2, Р.2.1.2, Р.2.2, Р.3, Р.4)
5. Сохраняем xlsx → вызываем `soffice --headless --convert-to pdf`
6. Читаем PDF и возвращаем bytes

Требуется LibreOffice в контейнере (Dockerfile).
"""
from __future__ import annotations

import os
import subprocess
import tempfile
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# Пути
# ============================================================

def _template_path() -> Path:
    here = Path(__file__).resolve().parent.parent.parent
    return here / "templates" / "knd_1152017" / "declaration_template_2024.xlsx"


# ============================================================
# Хелперы записи по клеткам
# ============================================================

def _col_letter(col_idx: int) -> str:
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(ord('A') + rem) + result
    return result


def _col_index(letter: str) -> int:
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _safe_set(ws: Worksheet, coord: str, value: Any) -> None:
    """Безопасная запись: если ячейка внутри merged range,
    пишет в его верхний-левый угол; иначе напрямую."""
    for merged_range in list(ws.merged_cells.ranges):
        if coord in merged_range:
            ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
            return
    ws[coord] = value


def _write_cell_row(
    ws: Worksheet,
    row: int,
    start_col_letter: str,
    n_cells: int,
    value: str,
    step: int = 2,
    align: str = "left",
) -> None:
    """Разносит строку value по n_cells ячейкам (шаг step колонок).
    Поддерживает merged cells."""
    start_idx = _col_index(start_col_letter)
    s = value or ""
    if align == "right":
        s = s.rjust(n_cells)[-n_cells:]
    elif align == "center":
        pad = max(0, (n_cells - len(s)) // 2)
        s = (" " * pad + s).ljust(n_cells)[:n_cells]
    else:
        s = s.ljust(n_cells)[:n_cells]

    for i, ch in enumerate(s):
        col = _col_letter(start_idx + i * step)
        coord = f"{col}{row}"
        _safe_set(ws, coord, ch if ch.strip() else None)


def _fmt_int(amount: Any) -> str:
    """Сумма → цифровая строка без копеек."""
    if amount is None or amount == "":
        return ""
    try:
        return str(int(Decimal(str(amount))))
    except Exception:
        return ""


# ============================================================
# Заполнение Титульного листа
# ============================================================

def _fill_title(ws: Worksheet, data: dict) -> None:
    """Заполняет Титульный лист данными налогоплательщика."""
    # ИНН 12 клеток Y1..AU1 шаг 2
    _write_cell_row(ws, 1, "Y", 12, data.get("inn", ""))
    # КПП 9 клеток Y4..AO4 шаг 2 (у ИП пусто)
    _write_cell_row(ws, 4, "Y", 9, data.get("kpp", ""))
    # Стр. 001 — дефолт в шаблоне

    # Номер корректировки X11 Z11 AB11
    _write_cell_row(ws, 11, "X", 3, str(data.get("correction_number", 0)))
    # Налоговый период 34
    _write_cell_row(ws, 11, "AQ", 2, "34")
    # Отчётный год BQ11..BW11
    _write_cell_row(ws, 11, "BQ", 4, str(data.get("year", "")))

    # ИФНС L13..R13
    _write_cell_row(ws, 13, "L", 4, data.get("ifns_code", ""))
    # По месту 120
    _write_cell_row(ws, 13, "BK", 3, "120")

    # ФИО 4 строки × 40 клеток
    fio_lines = _split_fio(data.get("fio", ""), max_per_line=40)
    for i, line_text in enumerate(fio_lines[:4]):
        row_num = 16 + i * 2
        _write_cell_row(ws, row_num, "A", 40, line_text)

    # Объект = 1
    _safe_set(ws, "Q29", "1")
    # Страниц = 4
    _write_cell_row(ws, 40, "C", 3, "4")
    # Подписант 1 = ИП сам
    _safe_set(ws, "J44", str(data.get("signer_kind", 1)))

    # Дата подписания S53/AB53/AH53
    sd = data.get("signing_date")
    if sd:
        date_str = sd.strftime("%d%m%Y") if hasattr(sd, "strftime") else ""
        _write_cell_row(ws, 53, "S", 2, date_str[:2])
        _write_cell_row(ws, 53, "AB", 2, date_str[2:4])
        _write_cell_row(ws, 53, "AH", 4, date_str[4:8])


def _compress_title_layout(ws: Worksheet) -> None:
    """Структурная модификация Титула: освобождает зону y=30..180pt
    внизу стр.1 под штамп ЭДО.

    row 30-39: 10 пустых строк × 17.25pt = 172.5pt → 5pt, экономим ~122pt
    row 55-58: 4 строки × 11pt = 44pt → 3pt, экономим ~32pt
    row 60-63: минимум 2pt
    Итого ~160pt свободного места внизу.
    """
    for row in range(30, 40):
        ws.row_dimensions[row].height = 5.0
    for row in range(55, 59):
        ws.row_dimensions[row].height = 3.0
    for row in range(60, 64):
        ws.row_dimensions[row].height = 2.0


def _split_fio(fio: str, max_per_line: int = 40) -> list[str]:
    words = (fio or "").upper().split()
    lines = []
    cur = ""
    for w in words:
        test = f"{cur} {w}".strip()
        if len(test) <= max_per_line:
            cur = test
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


# ============================================================
# Раздел 1.1
# ============================================================

def _fill_section_1_1(ws: Worksheet, data: dict) -> None:
    """ОКТМО + суммы авансов (координаты приблизительные — калибруются итеративно)."""
    oktmo = data.get("oktmo", "") or ""
    _write_cell_row(ws, 10, "M", 11, oktmo)
    amounts = [
        (14, _fmt_int(data.get("advance_q1"))),   # 020
        (18, _fmt_int(data.get("advance_h1"))),   # 040
        (22, _fmt_int(data.get("advance_h1_reduction"))),  # 050
        (26, _fmt_int(data.get("advance_9m"))),   # 070
        (30, _fmt_int(data.get("advance_9m_reduction"))),  # 080
        (34, _fmt_int(data.get("tax_year_payable"))),  # 100
        (38, _fmt_int(data.get("patent_offset"))),     # 101
        (42, _fmt_int(data.get("tax_year_reduction"))), # 110
    ]
    for row, val in amounts:
        _write_cell_row(ws, row, "M", 12, val, align="right")


# ============================================================
# Раздел 2.1.1
# ============================================================

def _fill_section_2_1_1(ws: Worksheet, data: dict) -> None:
    """Код ставки, доходы, ставки, исчисленный налог."""
    _safe_set(ws, "Q10", str(data.get("tax_rate_code", 1)))
    _safe_set(ws, "Q14", str(data.get("taxpayer_sign", 2)))

    incomes = [
        (20, _fmt_int(data.get("income_q1"))),
        (24, _fmt_int(data.get("income_h1"))),
        (28, _fmt_int(data.get("income_9m"))),
        (32, _fmt_int(data.get("income_y"))),
    ]
    for row, val in incomes:
        _write_cell_row(ws, row, "M", 12, val, align="right")

    rates = [
        (36, data.get("tax_rate_q1", Decimal("6.0"))),
        (38, data.get("tax_rate_h1", Decimal("6.0"))),
        (40, data.get("tax_rate_9m", Decimal("6.0"))),
        (42, data.get("tax_rate_y", Decimal("6.0"))),
    ]
    for row, rate in rates:
        try:
            d = Decimal(str(rate))
            int_part = str(int(d))
            dec_part = str(int((d - int(d)) * 10))
            _safe_set(ws, f"M{row}", int_part)
            _safe_set(ws, f"O{row}", dec_part)
        except Exception:
            pass

    calcs = [
        (48, _fmt_int(data.get("tax_calc_q1"))),
        (52, _fmt_int(data.get("tax_calc_h1"))),
        (56, _fmt_int(data.get("tax_calc_9m"))),
        (60, _fmt_int(data.get("tax_calc_y"))),
    ]
    for row, val in calcs:
        _write_cell_row(ws, row, "M", 12, val, align="right")


# ============================================================
# Раздел 2.1.1 продолжение
# ============================================================

def _fill_section_2_1_1_cont(ws: Worksheet, data: dict) -> None:
    contribs = [
        (10, _fmt_int(data.get("contrib_q1"))),
        (14, _fmt_int(data.get("contrib_h1"))),
        (18, _fmt_int(data.get("contrib_9m"))),
        (22, _fmt_int(data.get("contrib_y"))),
    ]
    for row, val in contribs:
        _write_cell_row(ws, row, "M", 12, val, align="right")


# ============================================================
# LibreOffice xlsx → pdf
# ============================================================

def _convert_xlsx_to_pdf(xlsx_path: Path, out_dir: Path) -> Path:
    cmd = [
        "soffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(out_dir),
        str(xlsx_path),
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    except FileNotFoundError:
        raise RuntimeError(
            "LibreOffice не установлен (soffice не найден). "
            "Добавьте libreoffice-core + libreoffice-calc в Dockerfile."
        )
    except subprocess.TimeoutExpired:
        raise RuntimeError("LibreOffice конвертация > 60s")

    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice failed (code {result.returncode}):\n"
            f"STDOUT: {result.stdout}\nSTDERR: {result.stderr}"
        )

    pdf_path = out_dir / (xlsx_path.stem + ".pdf")
    if not pdf_path.exists():
        raise RuntimeError(f"PDF не создан: {pdf_path}")
    return pdf_path


# ============================================================
# Public API
# ============================================================

def render_declaration_pdf(
    *,
    taxpayer,
    tax_period_year: int,
    tax_result,
    correction_number: int = 0,
    signing_date: datetime | None = None,
) -> bytes:
    """Рендер декларации через xlsx-шаблон + LibreOffice."""
    template = _template_path()
    if not template.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template}")

    data = dict(
        inn=taxpayer.inn,
        kpp="",
        fio=taxpayer.fio,
        oktmo=taxpayer.oktmo,
        ifns_code=taxpayer.ifns_code,
        year=tax_period_year,
        correction_number=correction_number,
        signing_date=signing_date or datetime.now(),
        signer_kind=1,
    )
    decl_data = getattr(tax_result, "decl_data", {}) or {}
    data.update(decl_data)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        work_xlsx = tmp / "declaration_filled.xlsx"

        wb = openpyxl.load_workbook(template)
        if "Титул" in wb.sheetnames:
            _fill_title(wb["Титул"], data)
            # Структурная модификация — освобождение зоны под штамп
            _compress_title_layout(wb["Титул"])
        if "Раздел 1.1" in wb.sheetnames:
            _fill_section_1_1(wb["Раздел 1.1"], data)
        if "Раздел 2.1.1" in wb.sheetnames:
            _fill_section_2_1_1(wb["Раздел 2.1.1"], data)
        if "Раздел 2.1.1 (продолжение)" in wb.sheetnames:
            _fill_section_2_1_1_cont(wb["Раздел 2.1.1 (продолжение)"], data)

        to_remove = [
            "Раздел 1.2",
            "Раздел 2.1.2",
            "Раздел 2.1.2 (продолжение)",
            "Раздел 2.2",
            "Раздел 2.2 (продолжение)",
            "Раздел 3",
            "Раздел 4",
        ]
        for name in to_remove:
            if name in wb.sheetnames:
                del wb[name]

        wb.save(work_xlsx)
        pdf_path = _convert_xlsx_to_pdf(work_xlsx, tmp)
        return pdf_path.read_bytes()
