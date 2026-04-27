"""
OFD (Оператор Фискальных Данных) receipt file parser.

Supports the standard ОФД xlsx export format with columns:
- РН, Место расчетов, Касса, Дата ФД, Тип ФД, Номер ФД,
  Признак расчета, Сумма чека, Наличные, Безналичные, Ошибки ФЛК

The parser:
1. Reads only "Кассовый чек" rows (skipping shift open/close reports).
2. Splits every receipt into its cash-part and card-part (a mixed receipt
   becomes two logical rows, one per payment type).
3. Marks "Приход" as sale and "Возврат прихода" as refund.

It also provides helpers to aggregate receipts by day/payment-type, which
is consumed by the revenue-reconciliation logic (OFD vs. bank acquiring).
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


# Columns we expect (tolerant mapping — case/spacing insensitive)
COL_ALIASES: Dict[str, Tuple[str, ...]] = {
    "rn":            ("рн", "rn"),
    "location":      ("место расчетов", "место расчётов"),
    "kkt":           ("касса", "номер ккт", "ккт"),
    "fd_date":       ("дата фд", "дата",),
    "fd_type":       ("тип фд",),
    "fd_number":     ("номер фд", "номер",),
    "operation":     ("признак расчета", "признак расчёта"),
    "total":         ("сумма чека", "итог", "итого"),
    "cash":          ("наличные",),
    "card":          ("безналичные", "электронные"),
    "errors":        ("ошибки флк", "ошибки"),
}


class OfdParser:
    """Parse ОФД xlsx receipt exports."""

    RECEIPT_TYPE = "Кассовый чек"

    def parse(self, file_path: str) -> Dict[str, Any]:
        """Parse an ОФД xlsx file.

        Returns:
            {
                "receipts": [
                    {
                        "receipt_date": datetime,
                        "amount": Decimal,
                        "payment_type": "cash"|"card",
                        "operation_type": "sale"|"refund",
                        "receipt_number": str,
                        "kkt_number": str,
                        "point_of_sale": str,
                    }, ...
                ],
                "total_receipts": int,
                "total_cash": Decimal,
                "total_card": Decimal,
                "total_refund_cash": Decimal,
                "total_refund_card": Decimal,
                "period_start": date|None,
                "period_end": date|None,
                "errors": [str],
                "warnings": [str],
            }
        """
        result: Dict[str, Any] = {
            "receipts": [],
            "total_receipts": 0,
            "total_cash": Decimal("0"),
            "total_card": Decimal("0"),
            "total_refund_cash": Decimal("0"),
            "total_refund_card": Decimal("0"),
            "period_start": None,
            "period_end": None,
            "errors": [],
            "warnings": [],
        }

        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            result["errors"].append(f"openpyxl is required: {exc}")
            return result

        path = Path(file_path)
        if not path.exists():
            result["errors"].append(f"File not found: {file_path}")
            return result

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            result["errors"].append(f"Cannot open xlsx: {exc}")
            return result

        # Pick the first sheet that has a header matching any expected column
        ws = None
        for sheet in wb.worksheets:
            header_row = self._find_header_row(sheet)
            if header_row is not None:
                ws = sheet
                break

        if ws is None:
            result["errors"].append("Подходящий лист не найден (нет колонки 'Тип ФД' или 'Сумма чека').")
            return result

        header_row_idx, col_map = self._build_col_map(ws)
        if "fd_type" not in col_map or "operation" not in col_map:
            result["errors"].append("В файле нет обязательных колонок 'Тип ФД' / 'Признак расчета'.")
            return result

        min_date: Optional[date] = None
        max_date: Optional[date] = None

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            fd_type = self._cell(row, col_map, "fd_type")
            if not fd_type or self.RECEIPT_TYPE not in str(fd_type):
                # Skip shift open/close reports and anything else
                continue

            operation = self._cell(row, col_map, "operation")
            op_type = self._classify_operation(operation)
            if op_type is None:
                # Unknown operation sign (correction, etc.) — skip
                result["warnings"].append(f"Пропущен чек с признаком '{operation}'")
                continue

            raw_date = self._cell(row, col_map, "fd_date")
            receipt_dt = self._parse_datetime(raw_date)
            if receipt_dt is None:
                result["warnings"].append(f"Не удалось распарсить дату: {raw_date}")
                continue

            d = receipt_dt.date()
            if min_date is None or d < min_date:
                min_date = d
            if max_date is None or d > max_date:
                max_date = d

            cash_amount = self._parse_decimal(self._cell(row, col_map, "cash")) or Decimal("0")
            card_amount = self._parse_decimal(self._cell(row, col_map, "card")) or Decimal("0")
            total_amount = self._parse_decimal(self._cell(row, col_map, "total")) or Decimal("0")

            # If cash/card columns are empty but the total is present, treat it as card-only
            if cash_amount == 0 and card_amount == 0 and total_amount > 0:
                card_amount = total_amount

            fd_number = str(self._cell(row, col_map, "fd_number") or "")
            kkt = str(self._cell(row, col_map, "kkt") or "")
            location = str(self._cell(row, col_map, "location") or "")

            # Emit one logical row per payment type present
            if cash_amount > 0:
                result["receipts"].append({
                    "receipt_date": receipt_dt,
                    "amount": cash_amount,
                    "payment_type": "cash",
                    "operation_type": op_type,
                    "receipt_number": fd_number,
                    "kkt_number": kkt,
                    "point_of_sale": location,
                })
                if op_type == "sale":
                    result["total_cash"] += cash_amount
                else:
                    result["total_refund_cash"] += cash_amount

            if card_amount > 0:
                result["receipts"].append({
                    "receipt_date": receipt_dt,
                    "amount": card_amount,
                    "payment_type": "card",
                    "operation_type": op_type,
                    "receipt_number": fd_number,
                    "kkt_number": kkt,
                    "point_of_sale": location,
                })
                if op_type == "sale":
                    result["total_card"] += card_amount
                else:
                    result["total_refund_card"] += card_amount

        result["total_receipts"] = len(result["receipts"])
        result["period_start"] = min_date
        result["period_end"] = max_date
        wb.close()
        return result

    # -------------------- helpers --------------------

    def _find_header_row(self, ws) -> Optional[int]:
        """Find the header row by looking for column aliases in the first 5 rows."""
        wanted = {alias for names in COL_ALIASES.values() for alias in names}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            normalized = [str(c).strip().lower() if c is not None else "" for c in row]
            if any(cell in wanted for cell in normalized):
                return row_idx
        return None

    def _build_col_map(self, ws) -> Tuple[int, Dict[str, int]]:
        """Return (header_row_index, {logical_name: 0-based_col_index})."""
        header_idx = self._find_header_row(ws) or 1
        header_row = next(ws.iter_rows(min_row=header_idx, max_row=header_idx, values_only=True))
        normalized = [str(c).strip().lower() if c is not None else "" for c in header_row]

        col_map: Dict[str, int] = {}
        for logical, aliases in COL_ALIASES.items():
            for i, cell in enumerate(normalized):
                if cell in aliases:
                    col_map[logical] = i
                    break
        return header_idx, col_map

    @staticmethod
    def _cell(row: Iterable, col_map: Dict[str, int], key: str):
        idx = col_map.get(key)
        if idx is None:
            return None
        row_list = list(row)
        if idx >= len(row_list):
            return None
        return row_list[idx]

    @staticmethod
    def _classify_operation(value) -> Optional[str]:
        """Map 'Признак расчета' to sale/refund."""
        if value is None:
            return None
        s = str(value).strip().lower()
        if "приход" == s or s == "приход":
            return "sale"
        if s.startswith("приход") and "возврат" not in s:
            return "sale"
        if "возврат прихода" in s or "возврат" in s:
            return "refund"
        return None

    @staticmethod
    def _parse_datetime(value) -> Optional[datetime]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        s = str(value).strip()
        if not s:
            return None
        for fmt in (
            "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    @staticmethod
    def _parse_decimal(value) -> Optional[Decimal]:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            try:
                return Decimal(str(value))
            except InvalidOperation:
                return None
        s = str(value).strip().replace("\xa0", "").replace(" ", "")
        if not s:
            return None
        s = s.replace(",", ".")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None


def parse_ofd_xlsx(file_path: str) -> Dict[str, Any]:
    """Convenience wrapper."""
    return OfdParser().parse(file_path)


# ---------------------------------------------------------------------------
# Reconciliation: compare OFD card totals with bank acquiring deposits per day
# ---------------------------------------------------------------------------

def aggregate_by_day(receipts: List[Dict[str, Any]]) -> Dict[date, Dict[str, Decimal]]:
    """Aggregate parsed OFD receipts by calendar day.

    Returns: { day: {"cash": Decimal, "card": Decimal,
                      "refund_cash": Decimal, "refund_card": Decimal} }
    """
    out: Dict[date, Dict[str, Decimal]] = defaultdict(lambda: {
        "cash": Decimal("0"),
        "card": Decimal("0"),
        "refund_cash": Decimal("0"),
        "refund_card": Decimal("0"),
    })
    for r in receipts:
        d = r["receipt_date"].date() if isinstance(r["receipt_date"], datetime) else r["receipt_date"]
        key = ("refund_" if r["operation_type"] == "refund" else "") + r["payment_type"]
        out[d][key] += r["amount"]
    return dict(out)


def reconcile_daily(
    ofd_by_day: Dict[date, Dict[str, Decimal]],
    bank_acquiring_by_day: Dict[date, Decimal],
) -> List[Dict[str, Any]]:
    """Merge daily OFD totals with bank acquiring deposits.

    For each day returns:
        {
          "date": date,
          "ofd_cash": Decimal,
          "ofd_card": Decimal,
          "ofd_refund_cash": Decimal,
          "ofd_refund_card": Decimal,
          "bank_acquiring": Decimal,
          "acquiring_diff": ofd_card - bank_acquiring,
          "cash_revenue": ofd_cash - ofd_refund_cash,
        }
    """
    days = sorted(set(ofd_by_day.keys()) | set(bank_acquiring_by_day.keys()))
    out: List[Dict[str, Any]] = []
    for d in days:
        o = ofd_by_day.get(d, {})
        ofd_cash = o.get("cash", Decimal("0"))
        ofd_card = o.get("card", Decimal("0"))
        ref_cash = o.get("refund_cash", Decimal("0"))
        ref_card = o.get("refund_card", Decimal("0"))
        bank = bank_acquiring_by_day.get(d, Decimal("0"))
        out.append({
            "date": d,
            "ofd_cash": ofd_cash,
            "ofd_card": ofd_card,
            "ofd_refund_cash": ref_cash,
            "ofd_refund_card": ref_card,
            "bank_acquiring": bank,
            "acquiring_diff": ofd_card - ref_card - bank,
            "cash_revenue": ofd_cash - ref_cash,
        })
    return out
