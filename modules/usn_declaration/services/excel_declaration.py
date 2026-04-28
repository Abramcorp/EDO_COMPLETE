"""
Fill Russian ИП УСН declaration (КНД 1152017) Excel template.

Supports two form versions:
  - 2024 form (приказ ФНС от 02.10.2024): 11 sheets, quarterly breakdown
  - 2025 form (приказ Минфина от 22.06.2009 №58н): 3 sheets, annual totals
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
import shutil

DATA_DIR = Path(__file__).resolve().parent.parent / 'data'  # EDO_COMPLETE: depth -1 vs upstream

TEMPLATES = {
    2024: DATA_DIR / 'declaration_template_2024.xlsx',
    2025: DATA_DIR / 'declaration_template_2025.xlsx',
}
# Fallback: old template for years without a specific template
TEMPLATE_XLSX = DATA_DIR / 'declaration_template.xlsx'


# =====================================================================
#  Helpers
# =====================================================================

_FONT = Font(name='Arial', size=11)


def _write_char(ws, coord, ch, font=_FONT):
    cell = ws[coord]
    cell.value = ch
    cell.font = font
    cell.alignment = Alignment(horizontal='center', vertical='center')


def write_chars(ws, row, cols, text, font=_FONT, align='left', pad_char=''):
    """Write each character of *text* into cells (row, cols[i])."""
    s = str(text) if text is not None else ''
    n = len(cols)
    if align == 'right':
        chars = list(s.rjust(n, pad_char) if pad_char else s.rjust(n))[-n:]
    else:
        chars = list(s)
        while len(chars) < n:
            chars.append(pad_char if pad_char else None)
    for i in range(n):
        ch = chars[i] if i < len(chars) else None
        if ch is None or ch == ' ':
            continue
        _write_char(ws, f'{cols[i]}{row}', ch, font)


def _remove_sheets(wb, keep):
    """Delete sheets not in *keep* set."""
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]


def _reorder_sheets(wb, order):
    """Reorder sheets to match *order* list."""
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            cur = wb.sheetnames.index(name)
            if cur != idx:
                wb.move_sheet(name, offset=idx - cur)


def _parse_date(date_val) -> tuple:
    """Return (dd, mm, yyyy) strings from date_val."""
    from datetime import date
    if isinstance(date_val, date):
        return date_val.strftime('%d'), date_val.strftime('%m'), date_val.strftime('%Y')
    s = str(date_val)
    dd, mm, yyyy = s.split('.')
    return dd, mm, yyyy



# =====================================================================
#  Legacy form (old template for ≤2023)
#  Kept for backward compatibility
# =====================================================================

# Old template cell maps (Титульный лист / Р.1.1 / Р.2.1.1 / Р.2.1.1 (продол.))
_OLD_TITUL_INN_COLS = ['M','N','O','P','Q','R','S','T','U','V','W','X']
_OLD_STR_COLS = ['X','Y','Z']

_OLD_R11 = {
    'oktmo_010': (12, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_020':  (14, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'oktmo_030': (17, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_040':  (19, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_050':  (22, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'oktmo_060': (25, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_070':  (27, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_080':  (30, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'oktmo_090': (33, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL']),
    'line_100':  (35, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_101':  (38, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_110':  (41, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
}

_OLD_R211 = {
    'line_101_pr': (12, ['AC']),
    'line_110': (17, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_111': (19, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_112': (21, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_113': (23, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_120_int': (27, ['AC']),  'line_120_dec': (27, ['AE']),
    'line_121_int': (29, ['AC']),  'line_121_dec': (29, ['AE']),
    'line_122_int': (31, ['AC']),  'line_122_dec': (31, ['AE']),
    'line_123_int': (33, ['AC']),  'line_123_dec': (33, ['AE']),
    'line_130': (40, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_131': (43, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_132': (46, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_133': (49, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
}

_OLD_R212 = {
    'line_140': (10, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_141': (14, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_142': (18, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
    'line_143': (22, ['AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM']),
}


# =====================================================================
#  New forms (2024+)
#  - 02.10.2024 № ЕД-7-3/813@ (XML version 5.08) — for tax period 2024
#  - 26.11.2025 № ЕД-7-3/1017@ (XML version 5.09) — for tax period 2025+
#  Both forms share the 11-sheet legacy structure (Титульный лист, Р.1.1, ...).
#  Difference: 5.08 has line_101 (rate marker 6%/8%) in Р.2.1.1; 5.09 does not.
# =====================================================================
_NEW_R211_5_08 = {
    # стр.101 — признак ставки (1=6%, 2=8% повышенная). Только в форме 5.08.
    'line_101_pr': (12, ['AC']),
    # стр.102 — признак налогоплательщика (1 или 2). В 5.08 на r=16, в 5.09 на r=12.
    'line_102_pr': (16, ['AC']),
    # стр.110-113 — доходы нарастающим итогом
    'line_110': (21, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_111': (23, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_112': (25, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_113': (27, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    # стр.120-123 — налоговая ставка (целая часть в AC, точка в AD, дробная в AE)
    'line_120_int': (31, ['AC']),  'line_120_dec': (31, ['AE']),
    'line_121_int': (33, ['AC']),  'line_121_dec': (33, ['AE']),
    'line_122_int': (35, ['AC']),  'line_122_dec': (35, ['AE']),
    'line_123_int': (37, ['AC']),  'line_123_dec': (37, ['AE']),
    # стр.130-133 — сумма исчисленного налога
    'line_130': (44, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_131': (48, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_132': (52, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
    'line_133': (56, ['AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']),
}


def _fill_form_5_08(template_path, out_path, project_data, decl_data):
    """Заполнить декларацию УСН по форме 02.10.2024 № ЕД-7-3/813@ (XML 5.08).
    Применяется для отчёта за налоговый период 2024 года.
    Главное отличие от 5.09: есть стр.101 в Разделе 2.1.1 (признак ставки 6%/8%)."""
    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    inn = str(project_data['inn']).zfill(12)

    # ---- Титульный лист ----
    ws = wb['Титульный лист']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '001')
    write_chars(ws, 10, ['H','I','J'], '0', pad_char='-')
    write_chars(ws, 10, ['W','X'], decl_data.get('period_code', '34'))
    write_chars(ws, 10, ['AK','AL','AM','AN'], str(project_data.get('tax_period_year', 2024)))
    write_chars(ws, 12, ['N','O','P','Q'], str(project_data.get('ifns_code', '')).zfill(4))
    write_chars(ws, 12, ['AL','AM','AN'], '120')
    fio_parts = project_data.get('fio', '').strip().split()
    fio_cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']
    for i, row in enumerate([14, 16, 18, 20]):
        val = fio_parts[i] if i < len(fio_parts) else ''
        if val:
            write_chars(ws, row, fio_cols, val)
    _write_char(ws, 'I28', '1')
    write_chars(ws, 39, ['C','D','E'], '004')
    _write_char(ws, 'B44', '1')
    dd, mm, yyyy = _parse_date(decl_data.get('date_presented', '27.01.2025'))
    write_chars(ws, 52, ['K','L'], dd)
    write_chars(ws, 52, ['N','O'], mm)
    write_chars(ws, 52, ['Q','R','S','T'], yyyy)

    # ---- Р.1.1 ---- (структура совпадает с _OLD_R11)
    ws = wb['Р.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '002')
    sec11 = decl_data.get('section_1_1', {})
    oktmo = str(project_data.get('oktmo', '')).ljust(11, '-')[:11]
    for key in ('oktmo_010','oktmo_030','oktmo_060','oktmo_090'):
        row, cols = _OLD_R11[key]
        write_chars(ws, row, cols, oktmo)
    for key in ('line_020','line_040','line_050','line_070','line_080',
                'line_100','line_101','line_110'):
        row, cols = _OLD_R11[key]
        val = sec11.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 ---- (свои координаты для 5.08, со стр.101 на r=12)
    ws = wb['Р.2.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '003')
    sec211 = decl_data.get('section_2_1_1', {})

    # стр.101 (признак ставки 6%/8%) — ТОЛЬКО в форме 5.08
    v101 = str(sec211.get('line_101_pr', '1'))
    row, cols = _NEW_R211_5_08['line_101_pr']
    _write_char(ws, f'{cols[0]}{row}', v101)
    # стр.102 (признак налогоплательщика)
    v102 = str(sec211.get('line_102_pr', '2'))
    row, cols = _NEW_R211_5_08['line_102_pr']
    _write_char(ws, f'{cols[0]}{row}', v102)

    for key in ('line_110','line_111','line_112','line_113'):
        row, cols = _NEW_R211_5_08[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')
    for key in ('line_120','line_121','line_122','line_123'):
        val = sec211.get(key, 6.0)
        int_part = str(int(val))
        dec_part = str(int(round((float(val) - int(val)) * 10)))
        ri, ci = _NEW_R211_5_08[f'{key}_int']
        rd, cd = _NEW_R211_5_08[f'{key}_dec']
        write_chars(ws, ri, ci, int_part)
        write_chars(ws, rd, cd, dec_part)
    for key in ('line_130','line_131','line_132','line_133'):
        row, cols = _NEW_R211_5_08[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 (продол.) ---- (структура совпадает с _OLD_R212)
    ws = wb['Р.2.1.1 (продол.)']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '004')
    for key in ('line_140','line_141','line_142','line_143'):
        row, cols = _OLD_R212[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    keep = {'Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'}
    _remove_sheets(wb, keep)
    _reorder_sheets(wb, ['Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'])

    wb.save(out_path)
    print(f'Saved 5.08 declaration: {out_path} (sheets: {wb.sheetnames})')


def _fill_form_5_09(template_path, out_path, project_data, decl_data):
    """Заполнить декларацию УСН по форме 26.11.2025 № ЕД-7-3/1017@ (XML 5.09).
    Применяется для отчёта за налоговый период 2025 года и далее.
    Координаты Р.2.1.1 совпадают с _OLD_R211 (стр.101 удалена, стр.102 на r=12)."""
    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    inn = str(project_data['inn']).zfill(12)

    # ---- Титульный лист ----
    ws = wb['Титульный лист']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '001')
    write_chars(ws, 10, ['H','I','J'], '0', pad_char='-')
    write_chars(ws, 10, ['W','X'], decl_data.get('period_code', '34'))
    write_chars(ws, 10, ['AK','AL','AM','AN'], str(project_data.get('tax_period_year', 2025)))
    write_chars(ws, 12, ['N','O','P','Q'], str(project_data.get('ifns_code', '')).zfill(4))
    write_chars(ws, 12, ['AL','AM','AN'], '120')
    fio_parts = project_data.get('fio', '').strip().split()
    fio_cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']
    for i, row in enumerate([14, 16, 18, 20]):
        val = fio_parts[i] if i < len(fio_parts) else ''
        if val:
            write_chars(ws, row, fio_cols, val)
    _write_char(ws, 'I28', '1')
    write_chars(ws, 39, ['C','D','E'], '004')
    _write_char(ws, 'B44', '1')
    dd, mm, yyyy = _parse_date(decl_data.get('date_presented', '27.04.2026'))
    write_chars(ws, 52, ['K','L'], dd)
    write_chars(ws, 52, ['N','O'], mm)
    write_chars(ws, 52, ['Q','R','S','T'], yyyy)

    # ---- Р.1.1 ---- (использует _OLD_R11 — координаты совпадают)
    ws = wb['Р.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '002')
    sec11 = decl_data.get('section_1_1', {})
    oktmo = str(project_data.get('oktmo', '')).ljust(11, '-')[:11]
    for key in ('oktmo_010','oktmo_030','oktmo_060','oktmo_090'):
        row, cols = _OLD_R11[key]
        write_chars(ws, row, cols, oktmo)
    for key in ('line_020','line_040','line_050','line_070','line_080',
                'line_100','line_101','line_110'):
        row, cols = _OLD_R11[key]
        val = sec11.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 ---- (использует _OLD_R211; стр.101 НЕТ в форме 5.09,
    # на r=12 находится стр.102 (признак налогоплательщика))
    ws = wb['Р.2.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '003')
    sec211 = decl_data.get('section_2_1_1', {})

    # На r=12 в форме 5.09 — стр.102 (признак НП). Берём из decl_data line_102_pr,
    # с fallback на line_101_pr для совместимости со старым forматом данных.
    v102 = str(sec211.get('line_102_pr', sec211.get('line_101_pr', '2')))
    _write_char(ws, 'AC12', v102)

    for key in ('line_110','line_111','line_112','line_113'):
        row, cols = _OLD_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')
    for key in ('line_120','line_121','line_122','line_123'):
        val = sec211.get(key, 6.0)
        int_part = str(int(val))
        dec_part = str(int(round((float(val) - int(val)) * 10)))
        ri, ci = _OLD_R211[f'{key}_int']
        rd, cd = _OLD_R211[f'{key}_dec']
        write_chars(ws, ri, ci, int_part)
        write_chars(ws, rd, cd, dec_part)
    for key in ('line_130','line_131','line_132','line_133'):
        row, cols = _OLD_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 (продол.) ----
    ws = wb['Р.2.1.1 (продол.)']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '004')
    for key in ('line_140','line_141','line_142','line_143'):
        row, cols = _OLD_R212[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    keep = {'Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'}
    _remove_sheets(wb, keep)
    _reorder_sheets(wb, ['Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'])

    wb.save(out_path)
    print(f'Saved 5.09 declaration: {out_path} (sheets: {wb.sheetnames})')



def _fill_old(template_path, out_path, project_data, decl_data):
    """Fill the old template (pre-2024)."""
    shutil.copy(template_path, out_path)
    wb = load_workbook(out_path)
    inn = str(project_data['inn']).zfill(12)

    # ---- Титульный лист ----
    ws = wb['Титульный лист']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '001')
    write_chars(ws, 10, ['H','I','J'], '0', pad_char='-')
    write_chars(ws, 10, ['W','X'], decl_data.get('period_code', '34'))
    write_chars(ws, 10, ['AK','AL','AM','AN'], str(project_data.get('tax_period_year', 2024)))
    write_chars(ws, 12, ['N','O','P','Q'], str(project_data.get('ifns_code', '')).zfill(4))
    write_chars(ws, 12, ['AL','AM','AN'], '120')
    fio_parts = project_data.get('fio', '').strip().split()
    fio_cols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T',
                'U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN']
    for i, row in enumerate([14, 16, 18, 20]):
        val = fio_parts[i] if i < len(fio_parts) else ''
        if val:
            write_chars(ws, row, fio_cols, val)
    _write_char(ws, 'I28', '1')
    write_chars(ws, 39, ['C','D','E'], '004')
    _write_char(ws, 'B44', '1')
    dd, mm, yyyy = _parse_date(decl_data.get('date_presented', '27.01.2025'))
    # pr30 fix: дата подписи — в зону НАЛОГОПЛАТЕЛЬЩИКА (row 52, cols K..T),
    # а не в зону ИФНС 'Дата представления декларации' (row 54, cols AE..AN).
    write_chars(ws, 52, ['K','L'], dd)
    write_chars(ws, 52, ['N','O'], mm)
    write_chars(ws, 52, ['Q','R','S','T'], yyyy)

    # ---- Р.1.1 ----
    ws = wb['Р.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '002')
    sec11 = decl_data.get('section_1_1', {})
    oktmo = str(project_data.get('oktmo', '')).ljust(11, '-')[:11]
    for key in ('oktmo_010','oktmo_030','oktmo_060','oktmo_090'):
        row, cols = _OLD_R11[key]
        write_chars(ws, row, cols, oktmo)
    for key in ('line_020','line_040','line_050','line_070','line_080',
                'line_100','line_101','line_110'):
        row, cols = _OLD_R11[key]
        val = sec11.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 ----
    ws = wb['Р.2.1.1']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '003')
    sec211 = decl_data.get('section_2_1_1', {})
    v101 = str(sec211.get('line_101', '2'))
    _write_char(ws, 'AC12', v101)
    for key in ('line_110','line_111','line_112','line_113'):
        row, cols = _OLD_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')
    for key in ('line_120','line_121','line_122','line_123'):
        val = sec211.get(key, 6.0)
        int_part = str(int(val))
        dec_part = str(int(round((float(val) - int(val)) * 10)))
        ri, ci = _OLD_R211[f'{key}_int']
        rd, cd = _OLD_R211[f'{key}_dec']
        write_chars(ws, ri, ci, int_part)
        write_chars(ws, rd, cd, dec_part)
    for key in ('line_130','line_131','line_132','line_133'):
        row, cols = _OLD_R211[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    # ---- Р.2.1.1 (продол.) ----
    ws = wb['Р.2.1.1 (продол.)']
    write_chars(ws, 1, _OLD_TITUL_INN_COLS, inn)
    write_chars(ws, 4, _OLD_STR_COLS, '004')
    for key in ('line_140','line_141','line_142','line_143'):
        row, cols = _OLD_R212[key]
        val = sec211.get(key, 0)
        if val:
            write_chars(ws, row, cols, str(int(val)), pad_char='-')

    keep = {'Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'}
    _remove_sheets(wb, keep)
    _reorder_sheets(wb, ['Титульный лист', 'Р.1.1', 'Р.2.1.1', 'Р.2.1.1 (продол.)'])

    wb.save(out_path)
    print(f'Saved old declaration: {out_path} (листы: {wb.sheetnames})')


# =====================================================================
#  Public API
# =====================================================================

def get_template_for_year(year: int) -> Path:
    """Return the correct template path for the given tax year."""
    if year in TEMPLATES and TEMPLATES[year].exists():
        return TEMPLATES[year]
    return TEMPLATE_XLSX


def fill_declaration(template_path: Path, out_path: Path,
                     project_data: dict, decl_data: dict):
    """Fill declaration using the correct form version based on template."""
    name = template_path.name

    if 'template_2024' in name:
        _fill_form_5_08(template_path, out_path, project_data, decl_data)
    elif 'template_2025' in name:
        _fill_form_5_09(template_path, out_path, project_data, decl_data)
    else:
        _fill_old(template_path, out_path, project_data, decl_data)
